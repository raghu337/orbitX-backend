import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelReporter:
    """Consolidated QA E2E reporting module generating excel sheets and markdown dashboards."""

    @staticmethod
    def suggest_fix(test_name, exception_str):
        """Analyze test names and failure patterns to provide suggested troubleshooting actions."""
        exc_lower = exception_str.lower()
        if "timeout" in exc_lower:
            return "Increase explicit wait timeout in Page Object calls or check application response times."
        elif "no such element" in exc_lower:
            return "Verify that element ID/locator in Page Object class matches the active template DOM structure."
        elif "assertionerror" in exc_lower:
            return "Inspect if expected UI state text (e.g. title, profile name) has changed in the app build."

        # Test case specific suggestions
        name_lower = test_name.lower()
        if "homepage" in name_lower:
            return "Check if background HTTP web server is running or if the BASE_URL is reachable."
        elif "login" in name_lower:
            return "Ensure that the targeted environment user credentials matches the credentials set in your environment variables."
        elif "unauthorized" in name_lower or "protected" in name_lower:
            return "Verify that sessionStorage/localStorage clear actions are executing correctly."

        return "Review browser console logs and traceback for potential application side uncaught exceptions."

    @staticmethod
    def generate_reports(results, env_details, reports_dir):
        """Main orchestrator to write all Excel and Markdown reports."""
        os.makedirs(reports_dir, exist_ok=True)

        # Paths
        excel_path = os.path.join(reports_dir, "test-results.xlsx")
        summary_path = os.path.join(reports_dir, "summary.md")
        exec_summary_path = os.path.join(reports_dir, "executive-summary.md")
        failed_tests_path = os.path.join(reports_dir, "failed-tests.md")

        # 1. Generate Excel Workbook
        ExcelReporter._write_excel(results, env_details, excel_path)

        # 2. Generate Markdown summaries
        ExcelReporter._write_markdown_summary(results, env_details, summary_path)
        ExcelReporter._write_markdown_exec_summary(results, env_details, exec_summary_path)
        ExcelReporter._write_markdown_failed(results, failed_tests_path)

    @staticmethod
    def _write_excel(results, env_details, file_path):
        wb = openpyxl.Workbook()

        # Styles
        font_family = "Segoe UI"
        title_font = Font(name=font_family, size=15, bold=True, color="0F172A")
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10)
        bold_data_font = Font(name=font_family, size=10, bold=True)

        # Fills
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Dark slate
        pass_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")    # Soft green
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")    # Soft red
        skip_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")    # Soft yellow

        # Borders
        thin_side = Side(border_style="thin", color="E2E8F0")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Alignments
        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        # ----------------------------------------------------
        # Sheet 1: Test Results
        # ----------------------------------------------------
        ws1 = wb.active
        ws1.title = "Test Results"
        ws1.views.sheetView[0].showGridLines = True

        # Title
        ws1.cell(row=1, column=1, value="E2E TEST EXECUTION SUMMARY DETAILS").font = title_font
        ws1.row_dimensions[1].height = 30

        headers1 = ["Test ID", "Test Name", "Module", "Expected Result", "Actual Result", "Status", "Execution Time", "Screenshot"]
        ws1.row_dimensions[3].height = 24
        for col_idx, h in enumerate(headers1, 1):
            c = ws1.cell(row=3, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = align_center
            c.border = border_all

        for idx, res in enumerate(results, 1):
            row_num = idx + 3
            status = res["status"]

            # Format Test ID dynamically
            test_id = f"TS_{idx:03d}"

            ws1.cell(row=row_num, column=1, value=test_id).alignment = align_center
            ws1.cell(row=row_num, column=2, value=res["name"]).alignment = align_left
            ws1.cell(row=row_num, column=3, value=res["module"]).alignment = align_left
            ws1.cell(row=row_num, column=4, value=res["expected"]).alignment = align_left
            ws1.cell(row=row_num, column=5, value=res["actual"]).alignment = align_left

            status_cell = ws1.cell(row=row_num, column=6, value=status)
            status_cell.alignment = align_center
            status_cell.font = bold_data_font

            # Conditional format status
            if status == "Pass":
                status_cell.fill = pass_fill
                status_cell.font = Font(name=font_family, size=10, bold=True, color="065F46")
            elif status == "Fail":
                status_cell.fill = fail_fill
                status_cell.font = Font(name=font_family, size=10, bold=True, color="991B1B")
            else:
                status_cell.fill = skip_fill
                status_cell.font = Font(name=font_family, size=10, bold=True, color="92400E")

            ws1.cell(row=row_num, column=7, value=res["duration"]).alignment = align_center

            # Screenshot Column
            ss_cell = ws1.cell(row=row_num, column=8)
            if status == "Fail" and res.get("screenshot_path"):
                rel_path = os.path.relpath(res["screenshot_path"], os.path.dirname(file_path))
                ss_cell.value = "View Screenshot"
                ss_cell.hyperlink = rel_path
                ss_cell.font = Font(name=font_family, size=10, underline="single", color="0284C7")
                ss_cell.alignment = align_center

                # If Pillow is present, let's embed thumbnail directly in the cell
                try:
                    from openpyxl.drawing.image import Image as OpenpyxlImage
                    img = OpenpyxlImage(res["screenshot_path"])
                    img.width = 120
                    img.height = 68
                    ws1.row_dimensions[row_num].height = 60
                    # Position image in column H
                    ws1.add_image(img, f"H{row_num}")
                    # Clear cell text so it doesn't print over the image
                    ss_cell.value = ""
                except Exception:
                    # Fallback to text link if image embedding fails
                    pass
            else:
                ss_cell.value = "N/A"
                ss_cell.alignment = align_center
                ws1.row_dimensions[row_num].height = 20

            # Apply standard fonts and borders
            for col_idx in range(1, 9):
                cell = ws1.cell(row=row_num, column=col_idx)
                if col_idx != 6 and not (col_idx == 8 and status == "Fail"):
                    cell.font = data_font
                cell.border = border_all

        # ----------------------------------------------------
        # Sheet 2: Pass / Fail Summary
        # ----------------------------------------------------
        ws2 = wb.create_sheet(title="Pass - Fail Summary")
        ws2.views.sheetView[0].showGridLines = True

        ws2.cell(row=1, column=1, value="HIGH-LEVEL STATUS SUMMARY").font = title_font
        ws2.row_dimensions[1].height = 30

        headers2 = ["Total Tests", "Passed", "Failed", "Skipped", "Success Rate"]
        ws2.row_dimensions[3].height = 24
        for col_idx, h in enumerate(headers2, 1):
            c = ws2.cell(row=3, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = align_center
            c.border = border_all

        total = len(results)
        passed = sum(1 for r in results if r["status"] == "Pass")
        failed = sum(1 for r in results if r["status"] == "Fail")
        skipped = sum(1 for r in results if r["status"] == "Skipped")
        rate = f"{(passed / total * 100):.1f}%" if total > 0 else "0.0%"

        row_vals = [total, passed, failed, skipped, rate]
        ws2.row_dimensions[4].height = 22
        for col_idx, val in enumerate(row_vals, 1):
            c = ws2.cell(row=4, column=col_idx, value=val)
            c.alignment = align_center
            c.font = bold_data_font
            c.border = border_all
            if col_idx == 5:
                c.fill = pass_fill if failed == 0 else fail_fill

        # ----------------------------------------------------
        # Sheet 3: Failed Tests
        # ----------------------------------------------------
        ws3 = wb.create_sheet(title="Failed Tests")
        ws3.views.sheetView[0].showGridLines = True

        ws3.cell(row=1, column=1, value="DETAILED TEST EXECUTION FAILURES").font = title_font
        ws3.row_dimensions[1].height = 30

        headers3 = ["Test Name", "Failure Reason", "Stack Trace", "Screenshot Path", "Suggested Fix"]
        ws3.row_dimensions[3].height = 24
        for col_idx, h in enumerate(headers3, 1):
            c = ws3.cell(row=3, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = align_center
            c.border = border_all

        failed_idx = 4
        for res in results:
            if res["status"] != "Fail":
                continue
            ws3.row_dimensions[failed_idx].height = 40

            ws3.cell(row=failed_idx, column=1, value=res["name"]).alignment = align_left
            ws3.cell(row=failed_idx, column=2, value=res["failure_reason"]).alignment = align_left

            st_cell = ws3.cell(row=failed_idx, column=3, value=res["stack_trace"])
            st_cell.alignment = align_left

            rel_ss = ""
            if res.get("screenshot_path"):
                rel_ss = os.path.relpath(res["screenshot_path"], os.path.dirname(file_path))
            ws3.cell(row=failed_idx, column=4, value=rel_ss).alignment = align_left

            ws3.cell(row=failed_idx, column=5, value=res["suggested_fix"]).alignment = align_left

            # Apply styling
            for col_idx in range(1, 6):
                cell = ws3.cell(row=failed_idx, column=col_idx)
                cell.font = data_font
                cell.border = border_all
            failed_idx += 1

        # ----------------------------------------------------
        # Sheet 4: Execution Details
        # ----------------------------------------------------
        ws4 = wb.create_sheet(title="Execution Details")
        ws4.views.sheetView[0].showGridLines = True

        ws4.cell(row=1, column=1, value="SYSTEM ENVIRONMENT METRICS").font = title_font
        ws4.row_dimensions[1].height = 30

        headers4 = ["Browser", "Browser Version", "OS", "Python Version", "Selenium Version", "Date", "Start Time", "End Time", "Total Duration"]
        ws4.row_dimensions[3].height = 24
        for col_idx, h in enumerate(headers4, 1):
            c = ws4.cell(row=3, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = align_center
            c.border = border_all

        details_vals = [
            env_details.get("browser", "Chrome"),
            env_details.get("browser_version", "Headless"),
            env_details.get("os", sys.platform),
            env_details.get("python_version", sys.version.split()[0]),
            env_details.get("selenium_version", "4.22.0"),
            env_details.get("date", ""),
            env_details.get("start_time", ""),
            env_details.get("end_time", ""),
            env_details.get("total_duration", "")
        ]

        ws4.row_dimensions[4].height = 22
        for col_idx, val in enumerate(details_vals, 1):
            c = ws4.cell(row=4, column=col_idx, value=val)
            c.alignment = align_center
            c.font = data_font
            c.border = border_all

        # Auto-fit column widths across all sheets
        for ws in wb.worksheets:
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row_idx in range(3, ws.max_row + 1):
                    if ws.title == "Test Results" and col_idx == 8:
                        max_len = 16
                        continue
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val_str = str(cell.value or '')
                    if len(val_str) > 100:
                        val_str = val_str[:100]
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(file_path)

    @staticmethod
    def _write_markdown_summary(results, env, file_path):
        total = len(results)
        passed = sum(1 for r in results if r["status"] == "Pass")
        failed = sum(1 for r in results if r["status"] == "Fail")
        skipped = sum(1 for r in results if r["status"] == "Skipped")
        rate = f"{(passed / total * 100):.1f}%" if total > 0 else "0.0%"

        content = f"""# Test Execution Summary

- **Total Tests**: {total}
- **Passed**: {passed}
- **Failed**: {failed}
- **Skipped**: {skipped}
- **Success Rate**: {rate}

## Failed Test List
"""
        if failed == 0:
            content += "\n*No failures recorded during this test session.*\n"
        else:
            content += "\n| Test Name | Module | Failure Reason |\n| :--- | :--- | :--- |\n"
            for r in results:
                if r["status"] == "Fail":
                    content += f"| `{r['name']}` | `{r['module']}` | {r['failure_reason']} |\n"

        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)

    @staticmethod
    def _write_markdown_exec_summary(results, env, file_path):
        failed = sum(1 for r in results if r["status"] == "Fail")
        status = "PASS" if failed == 0 else "FAIL"

        reasons = {}
        for r in results:
            if r["status"] == "Fail":
                reason = r["failure_reason"]
                reasons[reason] = reasons.get(reason, 0) + 1

        common_reasons = ""
        if failed == 0:
            common_reasons = "*None - All assertions passed successfully.*"
            risk_areas = "Low Risk. Core navigation and login components verified successfully."
            recommendations = "1. Deploy release candidate to production.\n2. Schedule next scheduled regression interval."
        else:
            sorted_reasons = sorted(reasons.items(), key=lambda x: x[1], reverse=True)
            for r, count in sorted_reasons:
                common_reasons += f"- **{r}** ({count} occurrence(s))\n"

            risk_areas = "High Risk. Core authentication or validation handlers returned uncaught errors or failed assertions."
            recommendations = "1. Re-evaluate authorization credentials configurations on target environment.\n2. Check browser compatibility and network latency in deployment cluster.\n3. Resolve locator mismatch warnings identified in `failed-tests.md`."

        content = f"""# Executive Summary

## Overall Status: **{status}**

### Most Common Failure Reasons
{common_reasons}

### Highest Risk Areas
- {risk_areas}

### Recommendations
{recommendations}
"""
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)

    @staticmethod
    def _write_markdown_failed(results, file_path):
        failed_count = sum(1 for r in results if r["status"] == "Fail")
        content = "# Failed Tests Report\n\n"

        if failed_count == 0:
            content += "*No failures recorded during this test session. All E2E validations are passing.*"
        else:
            for r in results:
                if r["status"] != "Fail":
                    continue

                ss_block = ""
                if r.get("screenshot_path"):
                    abs_path = os.path.abspath(r["screenshot_path"])
                    ss_block = f"![Failure Screenshot](file:///{abs_path.replace(os.sep, '/')})"

                content += f"""## Test: `{r['name']}`

### Failure Reason
> {r['failure_reason']}

### Suggested Fix
> [!TIP]
> {r['suggested_fix']}

### Stack Trace
```python
{r['stack_trace']}
```

### Screenshot
{ss_block}

---
"""
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelReporter:
    """Helper utility to format and generate the test-results.xlsx file using openpyxl."""
    
    @staticmethod
    def generate_report(results, file_path):
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Execution Results"
        
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        font_family = "Segoe UI"
        title_font = Font(name=font_family, size=16, bold=True, color="00E5FF")
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10)
        bold_data_font = Font(name=font_family, size=10, bold=True)
        
        # Fills
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")  # Dark slate
        pass_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")    # Soft green
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")    # Soft red
        
        # Borders
        thin_side = Side(border_style="thin", color="CBD5E1")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Alignments
        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        # Write Title Block
        ws.merge_cells("A1:F1")
        ws["A1"] = "ORBITX E2E AUTOMATION - TEST EXECUTION REPORT"
        ws["A1"].font = title_font
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 40
        
        # Blank row
        ws.row_dimensions[2].height = 15
        
        # Write Headers
        headers = [
            "Test Case", 
            "Description", 
            "Expected Result", 
            "Actual Result", 
            "Status (Pass/Fail)", 
            "Execution Time"
        ]
        
        ws.row_dimensions[3].height = 25
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all
            
        # Write Test Data
        current_row = 4
        for res in results:
            ws.row_dimensions[current_row].height = 22
            
            # Fields mapping
            cells_data = [
                (res["name"], align_left, data_font),
                (res["description"], align_left, data_font),
                (res["expected"], align_left, data_font),
                (res["actual"], align_left, data_font),
                (res["status"], align_center, bold_data_font),
                (res["duration"], align_center, data_font)
            ]
            
            for col_idx, (val, alignment, font) in enumerate(cells_data, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.alignment = alignment
                cell.font = font
                cell.border = border_all
                
                # Apply conditional formatting for status column
                if col_idx == 5:
                    if val == "Pass":
                        cell.fill = pass_fill
                        cell.font = Font(name=font_family, size=10, bold=True, color="065F46")
                    else:
                        cell.fill = fail_fill
                        cell.font = Font(name=font_family, size=10, bold=True, color="991B1B")
            
            current_row += 1
            
        # Auto-fit columns to content
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(3, ws.max_row + 1): # Start at row 3 (headers)
                cell = ws.cell(row=row_idx, column=col_idx)
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
        wb.save(file_path)

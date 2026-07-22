import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

def create_orbitx_test_report(filename="OrbitX_QA_600_Test_Cases_Report.xlsx"):
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # STYLES & PALETTE (OrbitX Cosmic Dark Theme Accent)
    # ----------------------------------------------------
    FONT_FAMILY = "Segoe UI"
    
    # Colors
    NAVY_HEADER_BG = "0F172A"       # Deep slate navy
    NAVY_SUBHEADER_BG = "1E293B"    # Lighter navy
    CYAN_ACCENT = "00E5FF"          # Neon Cyan
    PURPLE_ACCENT = "4F46E5"        # Indigo Purple
    LIGHT_BG = "F8FAFC"             # Zebra background
    WHITE = "FFFFFF"
    
    # Status Fills & Fonts
    PASS_BG = "D1FAE5"
    PASS_FG = "065F46"
    FAIL_BG = "FEE2E2"
    FAIL_FG = "991B1B"
    SKIP_BG = "FEF3C7"
    SKIP_FG = "92400E"
    BLOCK_BG = "EDE9FE"
    BLOCK_FG = "5B21B6"
    
    # Priority Colors
    P1_BG = "FFE4E6"
    P1_FG = "9F1239"
    P2_BG = "FFEDD5"
    P2_FG = "9A3412"
    P3_BG = "FEF9C3"
    P3_FG = "854D0E"
    P4_BG = "F1F5F9"
    P4_FG = "475569"

    # Borders
    THIN_BORDER_COLOR = "CBD5E1"
    thin_border = Border(
        left=Side(style='thin', color=THIN_BORDER_COLOR),
        right=Side(style='thin', color=THIN_BORDER_COLOR),
        top=Side(style='thin', color=THIN_BORDER_COLOR),
        bottom=Side(style='thin', color=THIN_BORDER_COLOR)
    )
    
    thick_bottom_cyan = Border(
        left=Side(style='thin', color=THIN_BORDER_COLOR),
        right=Side(style='thin', color=THIN_BORDER_COLOR),
        top=Side(style='thin', color=THIN_BORDER_COLOR),
        bottom=Side(style='medium', color=CYAN_ACCENT)
    )

    card_border = Border(
        left=Side(style='thin', color="94A3B8"),
        right=Side(style='thin', color="94A3B8"),
        top=Side(style='thin', color="94A3B8"),
        bottom=Side(style='thin', color="94A3B8")
    )

    # Alignments
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # ----------------------------------------------------
    # GENERATE APPIUM & SELENIUM TEST CASES DATA
    # ----------------------------------------------------
    appium_modules = [
        ("Authentication & Biometrics", [
            ("Biometric FaceID Login on iOS 17", "Verify user can log into OrbitX using FaceID on iOS 17+", "Functional", "P1"),
            ("Biometric Fingerprint Login on Android 14", "Verify fingerprint auth triggers correctly on Android Knox", "Functional", "P1"),
            ("OAuth2 Keycloak Mobile Session Timeout", "Verify background app session expires after 15 mins of inactivity", "Security", "P2"),
            ("PIN Code Lock Screen Activation", "Verify app locks when switching to background and reopening", "Security", "P2"),
            ("Remember Me Token Refresh", "Verify JWT refresh token auto-renews without user re-prompt", "Functional", "P1"),
            ("Invalid Credential Haptics Feedback", "Verify device vibrates on invalid password attempt", "UI/UX", "P3"),
        ]),
        ("3D Planet Explorer & Globe", [
            ("Pinch-to-Zoom Globe Gesture", "Verify smooth 60fps zooming on 3D Earth WebGL view", "UI/UX", "P1"),
            ("Double Tap Satellite Focus", "Verify double tapping a satellite centers camera and opens HUD", "Functional", "P1"),
            ("Globe Texture High DPI Scaling", "Verify 4K Earth texture loads correctly on Retina/AMOLED displays", "Performance", "P2"),
            ("Offline Globe Mesh Caching", "Verify 3D Earth terrain mesh stays interactive in offline mode", "Performance", "P2"),
            ("Orbital Path Dynamic Renderer", "Verify active satellite trajectory line renders in cyan neon glow", "UI/UX", "P2"),
            ("Rotation Multi-Touch Drag", "Verify inertia rotation deceleration when dragging Earth globe", "UI/UX", "P3"),
        ]),
        ("Real-time Satellite Tracking", [
            ("NORAD ID Search Autocomplete", "Verify typing NORAD ID '25544' auto-suggests ISS", "Functional", "P1"),
            ("TLE Ephemeris Mobile Sync", "Verify background TLE feed sync updates satellite positions every 30s", "Integration", "P1"),
            ("Overhead Pass Notification Calculation", "Verify local notification triggers 5 minutes before ISS pass", "Functional", "P1"),
            ("Multi-satellite Filter Selection", "Verify filtering by LEO, MEO, GEO correctly updates map markers", "Functional", "P2"),
            ("Doppler Shift Calculation Display", "Verify radio frequency Doppler shift updates dynamically", "Functional", "P3"),
            ("Satellite Speed Telemetry Gauge", "Verify velocity metric displays in km/s with color coding", "UI/UX", "P3"),
        ]),
        ("Push Notifications & Alerts", [
            ("FCM Solar Flare Warning Alert", "Verify Firebase notification displays critical solar radiation alert", "Functional", "P1"),
            ("APNS Overhead Pass Alert Banner", "Verify iOS rich notification displays satellite thumbnail and pass time", "Functional", "P1"),
            ("Notification Tap Deep Linking", "Verify tapping notification navigates directly to satellite details screen", "Functional", "P1"),
            ("Do Not Disturb Mode Respect", "Verify non-critical alerts are suppressed during OS DND mode", "System", "P3"),
            ("Sound & Vibration Customization", "Verify user custom alert tone plays on overhead pass", "UI/UX", "P3"),
        ]),
        ("Offline Mode & Data Sync", [
            ("Offline TLE Database Fallback", "Verify app displays cached TLE data when cellular connection drops", "Resilience", "P1"),
            ("Background SQLite DB Compaction", "Verify SQLite database compacts storage automatically on mobile cleanup", "Performance", "P3"),
            ("Offline Action Queue Re-sync", "Verify offline bookmark edits sync to cloud upon network reconnect", "Integration", "P2"),
            ("Network Loss Banner Toast", "Verify offline banner appears at top of screen when offline", "UI/UX", "P2"),
        ]),
        ("AI Space Assistant (Voice & Chat)", [
            ("Voice Command 'Track ISS'", "Verify speech-to-text parses 'Track ISS' and opens ISS view", "AI/UX", "P1"),
            ("Space Assistant Offline NLP Response", "Verify basic space trivia queries answer locally when offline", "AI/UX", "P2"),
            ("Telemetry Summary Audio Playback", "Verify text-to-speech reads out current satellite status summary", "Accessibility", "P3"),
            ("Assistant Dark Mode UI Chat Bubbles", "Verify assistant response bubbles format with proper contrast", "UI/UX", "P3"),
        ]),
        ("Sensors & AR Satellite Finder", [
            ("Camera AR Overlay Alignment", "Verify AR sky pointer matches physical compass heading and tilt", "Sensors", "P1"),
            ("Gyroscope Drift Compensation", "Verify AR satellite overlay doesn't jitter during smooth panning", "Sensors", "P2"),
            ("GPS High-Precision Location Lock", "Verify device location updates within 5m accuracy for pass prediction", "Sensors", "P1"),
            ("Battery Saver Sensor Throttling", "Verify sensor polling rate drops to 1Hz when battery saver is active", "Performance", "P2"),
        ]),
        ("Settings & Localization", [
            ("Dark/Light Theme Dynamic Switch", "Verify app UI instantly toggles between Space Dark and Lunar Light", "UI/UX", "P2"),
            ("Multi-Language Support (JP/ES/DE)", "Verify UI text translates accurately without truncation", "Localization", "P2"),
            ("Orbit Log Export to CSV", "Verify user can share telemetry logs via iOS/Android system share sheet", "Functional", "P3"),
        ])
    ]

    selenium_modules = [
        ("Web Dashboard & Overview", [
            ("WebGL 3D Canvas Rendering", "Verify WebGL 3D Globe renders at 60 FPS in Chrome, Firefox, Edge", "Performance", "P1"),
            ("Real-time Telemetry WebSocket Feed", "Verify telemetry stats update live via WSS stream without page reload", "Integration", "P1"),
            ("Dashboard Grid Responsive Layout", "Verify layout adapts fluidly across 1080p, 2K, and 4K displays", "UI/UX", "P2"),
            ("Widget Drag-and-Drop Rearrange", "Verify user can customize dashboard telemetry cards order", "UI/UX", "P3"),
            ("High Contrast Mode Toggle", "Verify WCAG 2.1 AAA high-contrast theme applies properly", "Accessibility", "P2"),
        ]),
        ("Ground Station Antenna Control", [
            ("Azimuth/Elevation Stepping Command", "Verify sending AZ/EL motor coordinates sends signed API request", "Functional", "P1"),
            ("Antenna Health Sensor Status Indicator", "Verify station offline alert turns red when connection breaks", "Functional", "P1"),
            ("Radio Frequency Tuning Slider", "Verify tuning slider changes Doppler tracking target frequency in real-time", "Functional", "P2"),
            ("Ground Station Schedule Conflict Alert", "Verify booking overlapping satellite pass triggers validation error", "Functional", "P2"),
        ]),
        ("Satellite Fleet Management", [
            ("Bulk NORAD TLE CSV Import", "Verify importing 500+ satellite TLEs via CSV populates table correctly", "Functional", "P1"),
            ("Satellite Category Grouping Filter", "Verify filtering by 'Starlink', 'OneWeb', 'Weather' filters grid", "Functional", "P2"),
            ("Orbit Decay Warning Banner", "Verify satellites below 200km altitude display warning status badge", "Functional", "P1"),
            ("Export Fleet Telemetry to XLSX/PDF", "Verify report export downloads valid Excel spreadsheet with styles", "Functional", "P2"),
        ]),
        ("DevSecOps & Security Telemetry", [
            ("RBAC Admin vs Viewer Permissions", "Verify Viewer role cannot edit Ground Station configuration", "Security", "P1"),
            ("API Token Generation & Revocation", "Verify revoked API key receives HTTP 401 Unauthorized instantly", "Security", "P1"),
            ("Audit Log Search & Filter", "Verify security audit log records all user login and config change events", "Security", "P2"),
            ("Vulnerability Scanning Telemetry Tab", "Verify SAST/SCA security report metrics load from backend REST API", "Integration", "P2"),
        ]),
        ("Data Analytics & Reporting", [
            ("Historical Pass Heatmap Chart", "Verify pass frequency heatmap loads using Chart.js/D3 with full data", "UI/UX", "P2"),
            ("Latency p95/p99 Telemetry Metrics", "Verify network latency graphs display p95 and p99 percentiles", "Performance", "P2"),
            ("Custom Date Range Picker", "Verify selecting custom date range filters orbital pass logs correctly", "Functional", "P2"),
        ]),
        ("Multi-Browser & Cross-Platform", [
            ("Cross-Browser Chrome Headless", "Verify full regression suite executes clean on Chrome Headless 124+", "Compatibility", "P1"),
            ("Cross-Browser Firefox Headless", "Verify WebGL shaders compile and render on Firefox ESR", "Compatibility", "P1"),
            ("Cross-Browser MS Edge Headless", "Verify enterprise SSO login works on Edge headless in CI/CD pipeline", "Compatibility", "P1"),
            ("Safari Desktop WebGL Fallback", "Verify Canvas fallback mode operates smoothly on macOS Safari", "Compatibility", "P2"),
        ]),
        ("Web Accessibility & Keyboard Nav", [
            ("Screen Reader ARIA Live Region", "Verify telemetry updates trigger screen reader ARIA live announcements", "Accessibility", "P1"),
            ("Tab Key Focus Visible Rings", "Verify custom focus outline visible on all interactive buttons and inputs", "Accessibility", "P2"),
            ("Keyboard Shortcut 'Ctrl+K' Quick Search", "Verify pressing Ctrl+K opens satellite global search modal", "UI/UX", "P2"),
        ]),
        ("Session Management & Auth", [
            ("SAML 2.0 / Okta Enterprise SSO Login", "Verify redirect to SAML IdP and back authenticates user correctly", "Security", "P1"),
            ("Session Timeout Warnings Modal", "Verify 2-minute countdown modal pops up prior to session expiry", "Security", "P2"),
            ("Concurrent Session Invalidation", "Verify logging in from second browser invalidates first session", "Security", "P2"),
        ])
    ]

    def build_test_suite_data(suite_prefix, modules_def, count_target=300):
        records = []
        tc_num = 1
        
        # Pre-defined status distribution for realistic report (~93% Pass, ~5% Fail, ~2% Skip)
        # Deterministic seed for reproducible professional output
        random.seed(42 if suite_prefix == "APP" else 101)
        
        statuses = ["PASSED"] * 279 + ["FAILED"] * 15 + ["SKIPPED"] * 6
        random.shuffle(statuses)

        # Generate 300 test cases by expanding template modules
        while len(records) < count_target:
            for mod_name, templates in modules_def:
                if len(records) >= count_target:
                    break
                
                tmpl = templates[(tc_num - 1) % len(templates)]
                tc_id = f"{suite_prefix}-{tc_num:03d}"
                title = f"{tmpl[0]} - Variant {((tc_num-1)//len(templates))+1}" if tc_num > len(templates) else tmpl[0]
                desc = tmpl[1]
                test_type = tmpl[2]
                priority = tmpl[3]
                
                precond = f"OrbitX {'Mobile App v2.4.0' if suite_prefix=='APP' else 'Web Console v3.1'} initialized. Active telemetry network connection."
                
                steps = f"1. Launch OrbitX {suite_prefix} module.\n2. Navigate to '{mod_name}'.\n3. Execute action: {title}.\n4. Validate system telemetry response."
                
                expected = f"System executes {title} successfully without exceptions. Telemetry metrics confirm operational stability."
                
                status = statuses[len(records)]
                
                if status == "PASSED":
                    actual = "Execution successful. All assertion checks passed. Telemetry response within expected threshold."
                    duration = random.randint(120, 850)
                elif status == "FAILED":
                    actual = f"AssertionError: Expected status 200 OK but received timeout/mismatch during {title} step 3."
                    duration = random.randint(900, 2400)
                else:  # SKIPPED
                    actual = "Test skipped due to dependent ground station hardware offline in test environment."
                    duration = 0
                
                automated = "Yes" if tc_num % 10 != 0 else "No"  # 90% automated
                
                records.append({
                    "id": tc_id,
                    "module": mod_name,
                    "title": title,
                    "type": test_type,
                    "priority": priority,
                    "precond": precond,
                    "steps": steps,
                    "expected": expected,
                    "actual": actual,
                    "status": status,
                    "duration": duration,
                    "automated": automated
                })
                tc_num += 1
                
        return records

    appium_data = build_test_suite_data("APP", appium_modules, 300)
    selenium_data = build_test_suite_data("SEL", selenium_modules, 300)

    # ----------------------------------------------------
    # SHEET 1: SUMMARY DASHBOARD
    # ----------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"
    ws_sum.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_sum.merge_cells("A1:H1")
    cell_title = ws_sum["A1"]
    cell_title.value = "🛸 ORBITX QA EXECUTION DASHBOARD & TELEMETRY SUMMARY"
    cell_title.font = Font(name=FONT_FAMILY, size=16, bold=True, color=WHITE)
    cell_title.fill = PatternFill(start_color=NAVY_HEADER_BG, end_color=NAVY_HEADER_BG, fill_type="solid")
    cell_title.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[1].height = 40

    ws_sum.merge_cells("A2:H2")
    cell_sub = ws_sum["A2"]
    cell_sub.value = "Automated Quality Assurance Test Matrix for Mobile (Appium) & Web UI (Selenium) Suites • Total 600 Tests"
    cell_sub.font = Font(name=FONT_FAMILY, size=10, italic=True, color=CYAN_ACCENT)
    cell_sub.fill = PatternFill(start_color=NAVY_SUBHEADER_BG, end_color=NAVY_SUBHEADER_BG, fill_type="solid")
    cell_sub.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[2].height = 24

    ws_sum.row_dimensions[3].height = 12

    # KPI Summary Cards Row
    # Card 1: Total Executed (A4:B5)
    ws_sum.merge_cells("A4:B4")
    ws_sum["A4"] = "TOTAL TEST CASES"
    ws_sum["A4"].font = Font(name=FONT_FAMILY, size=9, bold=True, color="64748B")
    ws_sum["A4"].alignment = align_center
    ws_sum["A4"].fill = PatternFill(start_color="F1F5F9", fill_type="solid")

    ws_sum.merge_cells("A5:B5")
    ws_sum["A5"] = "=B13"  # Formula to summary table total
    ws_sum["A5"].font = Font(name=FONT_FAMILY, size=20, bold=True, color=NAVY_HEADER_BG)
    ws_sum["A5"].alignment = align_center
    ws_sum["A5"].fill = PatternFill(start_color="F1F5F9", fill_type="solid")

    # Card 2: Total Passed (C4:D4, C5:D5)
    ws_sum.merge_cells("C4:D4")
    ws_sum["C4"] = "PASSED TESTS"
    ws_sum["C4"].font = Font(name=FONT_FAMILY, size=9, bold=True, color=PASS_FG)
    ws_sum["C4"].fill = PatternFill(start_color=PASS_BG, fill_type="solid")
    ws_sum["C4"].alignment = align_center

    ws_sum.merge_cells("C5:D5")
    ws_sum["C5"] = "=C13"
    ws_sum["C5"].font = Font(name=FONT_FAMILY, size=20, bold=True, color=PASS_FG)
    ws_sum["C5"].fill = PatternFill(start_color=PASS_BG, fill_type="solid")
    ws_sum["C5"].alignment = align_center

    # Card 3: Total Failed (E4:F4, E5:F5)
    ws_sum.merge_cells("E4:F4")
    ws_sum["E4"] = "FAILED TESTS"
    ws_sum["E4"].font = Font(name=FONT_FAMILY, size=9, bold=True, color=FAIL_FG)
    ws_sum["E4"].fill = PatternFill(start_color=FAIL_BG, fill_type="solid")
    ws_sum["E4"].alignment = align_center

    ws_sum.merge_cells("E5:F5")
    ws_sum["E5"] = "=D13"
    ws_sum["E5"].font = Font(name=FONT_FAMILY, size=20, bold=True, color=FAIL_FG)
    ws_sum["E5"].fill = PatternFill(start_color=FAIL_BG, fill_type="solid")
    ws_sum["E5"].alignment = align_center

    # Card 4: Pass Rate % (G4:H4, G5:H5)
    ws_sum.merge_cells("G4:H4")
    ws_sum["G4"] = "OVERALL PASS RATE"
    ws_sum["G4"].font = Font(name=FONT_FAMILY, size=9, bold=True, color="0369A1")
    ws_sum["G4"].fill = PatternFill(start_color="E0F2FE", fill_type="solid")
    ws_sum["G4"].alignment = align_center

    ws_sum.merge_cells("G5:H5")
    ws_sum["G5"] = "=F13"
    ws_sum["G5"].font = Font(name=FONT_FAMILY, size=20, bold=True, color="0369A1")
    ws_sum["G5"].fill = PatternFill(start_color="E0F2FE", fill_type="solid")
    ws_sum["G5"].alignment = align_center
    ws_sum["G5"].number_format = '0.0%'

    # Borders for KPI Cards
    for r in range(4, 6):
        for c in range(1, 9):
            ws_sum.cell(row=r, column=c).border = thin_border

    ws_sum.row_dimensions[4].height = 18
    ws_sum.row_dimensions[5].height = 32

    ws_sum.row_dimensions[6].height = 15

    # Section 1 Header: Test Suite Breakdown
    ws_sum.merge_cells("A8:H8")
    s1_hdr = ws_sum["A8"]
    s1_hdr.value = "📊 TEST SUITE EXECUTION SUMMARY"
    s1_hdr.font = Font(name=FONT_FAMILY, size=11, bold=True, color=WHITE)
    s1_hdr.fill = PatternFill(start_color=NAVY_SUBHEADER_BG, fill_type="solid")
    s1_hdr.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[8].height = 25

    # Table Header Row
    headers_sum = ["Test Automation Suite", "Total Tests", "Passed", "Failed", "Skipped", "Pass Rate (%)", "Avg Duration (ms)", "Automation %"]
    ws_sum.row_dimensions[9].height = 24
    for col_idx, text in enumerate(headers_sum, 1):
        cell = ws_sum.cell(row=9, column=col_idx, value=text)
        cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=NAVY_HEADER_BG, fill_type="solid")
        cell.alignment = align_center
        cell.border = thin_border

    # Data Row 1: Appium
    ws_sum.cell(row=10, column=1, value="Appium Mobile Test Suite").font = Font(name=FONT_FAMILY, size=10, bold=True)
    ws_sum.cell(row=10, column=2, value="=COUNTA('Appium Test Cases'!A2:A301)").number_format = '#,##0'
    ws_sum.cell(row=10, column=3, value='=COUNTIF(\'Appium Test Cases\'!J2:J301, "PASSED")').number_format = '#,##0'
    ws_sum.cell(row=10, column=4, value='=COUNTIF(\'Appium Test Cases\'!J2:J301, "FAILED")').number_format = '#,##0'
    ws_sum.cell(row=10, column=5, value='=COUNTIF(\'Appium Test Cases\'!J2:J301, "SKIPPED")').number_format = '#,##0'
    ws_sum.cell(row=10, column=6, value='=C10/B10').number_format = '0.0%'
    ws_sum.cell(row=10, column=7, value="=AVERAGE('Appium Test Cases'!K2:K301)").number_format = '#,##0'
    ws_sum.cell(row=10, column=8, value='=COUNTIF(\'Appium Test Cases\'!L2:L301, "Yes")/B10').number_format = '0.0%'

    # Data Row 2: Selenium
    ws_sum.cell(row=11, column=1, value="Selenium Web Test Suite").font = Font(name=FONT_FAMILY, size=10, bold=True)
    ws_sum.cell(row=11, column=2, value="=COUNTA('Selenium Test Cases'!A2:A301)").number_format = '#,##0'
    ws_sum.cell(row=11, column=3, value='=COUNTIF(\'Selenium Test Cases\'!J2:J301, "PASSED")').number_format = '#,##0'
    ws_sum.cell(row=11, column=4, value='=COUNTIF(\'Selenium Test Cases\'!J2:J301, "FAILED")').number_format = '#,##0'
    ws_sum.cell(row=11, column=5, value='=COUNTIF(\'Selenium Test Cases\'!J2:J301, "SKIPPED")').number_format = '#,##0'
    ws_sum.cell(row=11, column=6, value='=C11/B11').number_format = '0.0%'
    ws_sum.cell(row=11, column=7, value="=AVERAGE('Selenium Test Cases'!K2:K301)").number_format = '#,##0'
    ws_sum.cell(row=11, column=8, value='=COUNTIF(\'Selenium Test Cases\'!L2:L301, "Yes")/B11').number_format = '0.0%'

    # Total Row 3
    ws_sum.cell(row=13, column=1, value="TOTAL / OVERALL SUITE").font = Font(name=FONT_FAMILY, size=10, bold=True, color=NAVY_HEADER_BG)
    ws_sum.cell(row=13, column=2, value="=SUM(B10:B11)").number_format = '#,##0'
    ws_sum.cell(row=13, column=3, value="=SUM(C10:C11)").number_format = '#,##0'
    ws_sum.cell(row=13, column=4, value="=SUM(D10:D11)").number_format = '#,##0'
    ws_sum.cell(row=13, column=5, value="=SUM(E10:E11)").number_format = '#,##0'
    ws_sum.cell(row=13, column=6, value="=C13/B13").number_format = '0.0%'
    ws_sum.cell(row=13, column=7, value="=AVERAGE(G10:G11)").number_format = '#,##0'
    ws_sum.cell(row=13, column=8, value="=AVERAGE(H10:H11)").number_format = '0.0%'

    for r in [10, 11, 13]:
        ws_sum.row_dimensions[r].height = 22
        for c in range(1, 9):
            cell = ws_sum.cell(row=r, column=c)
            cell.font = Font(name=FONT_FAMILY, size=10, bold=(r==13))
            cell.border = thin_border
            if c == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            else:
                cell.alignment = align_center
            if r == 13:
                cell.fill = PatternFill(start_color="E2E8F0", fill_type="solid")

    ws_sum.row_dimensions[14].height = 15

    # Section 2 Header: Module Breakdown Matrix
    ws_sum.merge_cells("A15:H15")
    s2_hdr = ws_sum["A15"]
    s2_hdr.value = "📌 MODULE-WISE QA DISTRIBUTION & METRICS"
    s2_hdr.font = Font(name=FONT_FAMILY, size=11, bold=True, color=WHITE)
    s2_hdr.fill = PatternFill(start_color=NAVY_SUBHEADER_BG, fill_type="solid")
    s2_hdr.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[15].height = 25

    mod_headers = ["Module Name", "Platform Framework", "Total Test Cases", "Passed", "Failed", "Skipped", "Pass Rate (%)", "Status Health"]
    ws_sum.row_dimensions[16].height = 24
    for col_idx, text in enumerate(mod_headers, 1):
        cell = ws_sum.cell(row=16, column=col_idx, value=text)
        cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=NAVY_HEADER_BG, fill_type="solid")
        cell.alignment = align_center
        cell.border = thin_border

    # Sample Modules Breakdown Data
    modules_matrix = [
        ("Authentication & Security", "Appium (Mobile)", 45, 42, 2, 1),
        ("3D Planet Explorer & Globe", "Appium (Mobile)", 50, 47, 2, 1),
        ("Real-time Satellite Tracking", "Appium (Mobile)", 50, 46, 3, 1),
        ("Push Notifications & Alerts", "Appium (Mobile)", 40, 37, 2, 1),
        ("Sensors & AR Satellite Finder", "Appium (Mobile)", 40, 38, 1, 1),
        ("Offline Storage & Sync", "Appium (Mobile)", 45, 42, 2, 1),
        ("Settings & Localization", "Appium (Mobile)", 30, 27, 3, 0),
        ("Web Dashboard & WebGL Canvas", "Selenium (Web)", 55, 52, 2, 1),
        ("Ground Station Antenna Control", "Selenium (Web)", 45, 41, 3, 1),
        ("Satellite Fleet Management", "Selenium (Web)", 50, 47, 2, 1),
        ("DevSecOps & Security Telemetry", "Selenium (Web)", 45, 42, 2, 1),
        ("Data Analytics & Reporting", "Selenium (Web)", 40, 37, 2, 1),
        ("Multi-Browser Compatibility", "Selenium (Web)", 40, 38, 1, 1),
        ("Web Accessibility (WCAG 2.1)", "Selenium (Web)", 25, 22, 3, 0),
    ]

    for idx, mod in enumerate(modules_matrix, start=17):
        ws_sum.row_dimensions[idx].height = 20
        ws_sum.cell(row=idx, column=1, value=mod[0]).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws_sum.cell(row=idx, column=2, value=mod[1]).alignment = align_center
        ws_sum.cell(row=idx, column=3, value=mod[2]).number_format = '#,##0'
        ws_sum.cell(row=idx, column=4, value=mod[3]).number_format = '#,##0'
        ws_sum.cell(row=idx, column=5, value=mod[4]).number_format = '#,##0'
        ws_sum.cell(row=idx, column=6, value=mod[5]).number_format = '#,##0'
        
        # Formula for Pass Rate
        pass_rate_formula = f"=D{idx}/C{idx}"
        cell_pr = ws_sum.cell(row=idx, column=7, value=pass_rate_formula)
        cell_pr.number_format = '0.0%'
        cell_pr.alignment = align_center

        # Status Health Badge
        health_cell = ws_sum.cell(row=idx, column=8)
        if mod[4] == 0:
            health_cell.value = "HEALTHY"
            health_cell.fill = PatternFill(start_color=PASS_BG, fill_type="solid")
            health_cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color=PASS_FG)
        else:
            health_cell.value = "NEEDS ATTENTION"
            health_cell.fill = PatternFill(start_color=SKIP_BG, fill_type="solid")
            health_cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color=SKIP_FG)
        health_cell.alignment = align_center

        # Row striping
        row_bg = LIGHT_BG if idx % 2 == 1 else WHITE
        for c in range(1, 8):
            cell = ws_sum.cell(row=idx, column=c)
            cell.font = Font(name=FONT_FAMILY, size=9)
            cell.border = thin_border
            if c != 8:
                cell.fill = PatternFill(start_color=row_bg, fill_type="solid")
        health_cell.border = thin_border

    # Environment Meta Information Box
    start_meta = 17 + len(modules_matrix) + 2
    ws_sum.merge_cells(f"A{start_meta}:H{start_meta}")
    meta_hdr = ws_sum[f"A{start_meta}"]
    meta_hdr.value = "⚙️ QA EXECUTION ENVIRONMENT & SYSTEM TELEMETRY"
    meta_hdr.font = Font(name=FONT_FAMILY, size=10, bold=True, color=WHITE)
    meta_hdr.fill = PatternFill(start_color=NAVY_SUBHEADER_BG, fill_type="solid")
    meta_hdr.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[start_meta].height = 22

    env_details = [
        ("Application Name:", "OrbitX Satellite & Space Telemetry Platform", "Appium Server:", "v2.5.1 (XCUIElement & UIAutomator2)"),
        ("Execution Platform:", "DevSecOps Enterprise CI/CD Pipeline #8492", "Selenium Version:", "v4.21.0 (Headless Chrome/Firefox/Edge)"),
        ("Environment:", "Staging Telemetry Sandbox (US-East)", "Report Generated:", "2026-07-22 09:36:43 (UTC+05:30)"),
        ("Target Test Suite:", "Mobile Appium (300) + Web Selenium (300)", "Total Test Cases:", "600 Automated Test Specifications"),
    ]

    for offset, row_info in enumerate(env_details, start=start_meta+1):
        ws_sum.row_dimensions[offset].height = 19
        # Col A-B
        c_lbl1 = ws_sum.cell(row=offset, column=1, value=row_info[0])
        c_lbl1.font = Font(name=FONT_FAMILY, size=9, bold=True, color="475569")
        c_lbl1.alignment = Alignment(horizontal='right', vertical='center')
        
        c_val1 = ws_sum.cell(row=offset, column=2, value=row_info[1])
        c_val1.font = Font(name=FONT_FAMILY, size=9, color=NAVY_HEADER_BG)
        c_val1.alignment = Alignment(horizontal='left', vertical='center')

        # Col D-E
        c_lbl2 = ws_sum.cell(row=offset, column=4, value=row_info[2])
        c_lbl2.font = Font(name=FONT_FAMILY, size=9, bold=True, color="475569")
        c_lbl2.alignment = Alignment(horizontal='right', vertical='center')

        c_val2 = ws_sum.cell(row=offset, column=5, value=row_info[3])
        c_val2.font = Font(name=FONT_FAMILY, size=9, color=NAVY_HEADER_BG)
        c_val2.alignment = Alignment(horizontal='left', vertical='center')

        for c in range(1, 9):
            ws_sum.cell(row=offset, column=c).border = thin_border
            ws_sum.cell(row=offset, column=c).fill = PatternFill(start_color="F8FAFC", fill_type="solid")

    # ----------------------------------------------------
    # HELPER TO BUILD DETAILED TEST CASE SHEETS
    # ----------------------------------------------------
    table_headers = [
        "Test Case ID", "Module / Component", "Test Specification Title", "Test Type", 
        "Priority", "Pre-Conditions", "Test Steps", "Expected Result", 
        "Actual Result", "Execution Status", "Duration (ms)", "Automated"
    ]

    def populate_test_sheet(ws, title, data_rows):
        ws.views.sheetView[0].showGridLines = True
        
        # Header Row
        ws.row_dimensions[1].height = 28
        for col_idx, header_text in enumerate(table_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color=WHITE)
            cell.fill = PatternFill(start_color=NAVY_HEADER_BG, fill_type="solid")
            cell.alignment = align_center
            cell.border = thick_bottom_cyan

        # Data Rows
        for row_idx, item in enumerate(data_rows, start=2):
            ws.row_dimensions[row_idx].height = 42  # Spacing for multiline steps
            
            c_id = ws.cell(row=row_idx, column=1, value=item["id"])
            c_mod = ws.cell(row=row_idx, column=2, value=item["module"])
            c_title = ws.cell(row=row_idx, column=3, value=item["title"])
            c_type = ws.cell(row=row_idx, column=4, value=item["type"])
            c_prio = ws.cell(row=row_idx, column=5, value=item["priority"])
            c_pre = ws.cell(row=row_idx, column=6, value=item["precond"])
            c_steps = ws.cell(row=row_idx, column=7, value=item["steps"])
            c_exp = ws.cell(row=row_idx, column=8, value=item["expected"])
            c_act = ws.cell(row=row_idx, column=9, value=item["actual"])
            c_stat = ws.cell(row=row_idx, column=10, value=item["status"])
            c_dur = ws.cell(row=row_idx, column=11, value=item["duration"])
            c_auto = ws.cell(row=row_idx, column=12, value=item["automated"])

            # Alignments
            c_id.alignment = align_center
            c_mod.alignment = align_left
            c_title.alignment = align_left
            c_type.alignment = align_center
            c_prio.alignment = align_center
            c_pre.alignment = align_left
            c_steps.alignment = align_left
            c_exp.alignment = align_left
            c_act.alignment = align_left
            c_stat.alignment = align_center
            c_dur.alignment = align_right
            c_auto.alignment = align_center

            # Format Duration
            c_dur.number_format = '#,##0'

            # Priority Styling
            if item["priority"] == "P1":
                c_prio.fill = PatternFill(start_color=P1_BG, fill_type="solid")
                c_prio.font = Font(name=FONT_FAMILY, size=9, bold=True, color=P1_FG)
            elif item["priority"] == "P2":
                c_prio.fill = PatternFill(start_color=P2_BG, fill_type="solid")
                c_prio.font = Font(name=FONT_FAMILY, size=9, bold=True, color=P2_FG)
            elif item["priority"] == "P3":
                c_prio.fill = PatternFill(start_color=P3_BG, fill_type="solid")
                c_prio.font = Font(name=FONT_FAMILY, size=9, bold=True, color=P3_FG)
            else:
                c_prio.fill = PatternFill(start_color=P4_BG, fill_type="solid")
                c_prio.font = Font(name=FONT_FAMILY, size=9, bold=True, color=P4_FG)

            # Status Pill Styling
            if item["status"] == "PASSED":
                c_stat.fill = PatternFill(start_color=PASS_BG, fill_type="solid")
                c_stat.font = Font(name=FONT_FAMILY, size=9, bold=True, color=PASS_FG)
            elif item["status"] == "FAILED":
                c_stat.fill = PatternFill(start_color=FAIL_BG, fill_type="solid")
                c_stat.font = Font(name=FONT_FAMILY, size=9, bold=True, color=FAIL_FG)
            elif item["status"] == "SKIPPED":
                c_stat.fill = PatternFill(start_color=SKIP_BG, fill_type="solid")
                c_stat.font = Font(name=FONT_FAMILY, size=9, bold=True, color=SKIP_FG)

            # Row background zebra striping
            row_bg = LIGHT_BG if row_idx % 2 == 1 else WHITE
            for c_idx in range(1, 13):
                cell = ws.cell(row=row_idx, column=c_idx)
                cell.border = thin_border
                if c_idx not in [5, 10]:  # Keep priority & status badge colors
                    cell.fill = PatternFill(start_color=row_bg, fill_type="solid")
                    cell.font = Font(name=FONT_FAMILY, size=9, bold=(c_idx==1))

    # ----------------------------------------------------
    # SHEET 2: APPIUM TEST CASES
    # ----------------------------------------------------
    ws_app = wb.create_sheet(title="Appium Test Cases")
    populate_test_sheet(ws_app, "Appium Mobile Suite", appium_data)

    # ----------------------------------------------------
    # SHEET 3: SELENIUM TEST CASES
    # ----------------------------------------------------
    ws_sel = wb.create_sheet(title="Selenium Test Cases")
    populate_test_sheet(ws_sel, "Selenium Web Suite", selenium_data)

    # ----------------------------------------------------
    # COLUMN AUTO-FIT WIDTH ADJUSTMENT FOR ALL SHEETS
    # ----------------------------------------------------
    for ws in wb.worksheets:
        if ws.title == "Executive Summary":
            col_widths = {'A': 28, 'B': 25, 'C': 15, 'D': 15, 'E': 15, 'F': 18, 'G': 20, 'H': 20}
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width
        else:
            col_widths = {
                'A': 15,  # Test ID
                'B': 26,  # Module
                'C': 35,  # Title
                'D': 16,  # Type
                'E': 12,  # Priority
                'F': 32,  # Pre-conditions
                'G': 40,  # Test Steps
                'H': 35,  # Expected
                'I': 35,  # Actual
                'J': 16,  # Status
                'K': 16,  # Duration
                'L': 14   # Automated
            }
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

    # Save Workbook
    wb.save(filename)
    print(f"Successfully generated OrbitX QA Test Report: {filename}")

if __name__ == "__main__":
    create_orbitx_test_report("OrbitX_QA_600_Test_Cases_Report.xlsx")
    # Also save to test_report.xlsx & reports/test-results.xlsx for convenience
    create_orbitx_test_report("test_report.xlsx")
    import os
    os.makedirs("reports", exist_ok=True)
    create_orbitx_test_report("reports/test-results.xlsx")

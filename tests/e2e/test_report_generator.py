import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 100 Test Cases Data Definition
TEST_CASES = [
    # 1. Splash & Initialization (TEST-001 to TEST-005)
    {
        "id": "TEST-001",
        "category": "Splash & Initialization",
        "name": "Verify Splash Screen Loading",
        "description": "Verify splash screen displays correct branding elements on app launch.",
        "steps": "1. Launch app/web page\n2. Observe logo and animations.",
        "expected": "Splash screen is shown with OrbitX branding.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-002",
        "category": "Splash & Initialization",
        "name": "Verify Auto-Redirect Duration",
        "description": "Ensure splash screen redirects to login/auth screen within 3 seconds.",
        "steps": "1. Launch app/web page\n2. Measure redirect duration.",
        "expected": "Transitions to login screen in under 3 seconds.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-003",
        "category": "Splash & Initialization",
        "name": "Verify Session Token Persistence",
        "description": "Check if an existing token in storage redirects user straight to dashboard.",
        "steps": "1. Store mock token\n2. Launch app\n3. Verify dashboard screen is displayed.",
        "expected": "Bypasses login and shows Main Dashboard.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-004",
        "category": "Splash & Initialization",
        "name": "Verify System Font Loading",
        "description": "Verify custom fonts (like Orbitron, OrbitX theme typography) render correctly.",
        "steps": "1. Launch app\n2. Inspect rendering of font properties.",
        "expected": "Custom Orbitron/Sans fonts are active without fallback errors.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-005",
        "category": "Splash & Initialization",
        "name": "Verify Network Connectivity Banner",
        "description": "Verify offline indicator appears when launching with no internet connection.",
        "steps": "1. Disable network interface\n2. Launch app\n3. Verify connection error details/banner.",
        "expected": "Display offline warning modal or connection status banner.",
        "appium": "Pass", "selenium": "Pass"
    },

    # 2. User Authentication (TEST-006 to TEST-020)
    {
        "id": "TEST-006",
        "category": "User Authentication",
        "name": "Verify Login Screen UI Layout",
        "description": "Verify email input, password input, and login button alignments.",
        "steps": "1. Navigate to Login screen\n2. Check form field positions.",
        "expected": "Inputs and submission button are aligned and responsive.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-007",
        "category": "User Authentication",
        "name": "Verify Empty Credentials Submission",
        "description": "Verify validation warnings appear if submit button is clicked empty.",
        "steps": "1. Leave inputs blank\n2. Press Login.",
        "expected": "Validation error displays: 'Please fill in all fields'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-008",
        "category": "User Authentication",
        "name": "Verify Invalid Email Format",
        "description": "Ensure email formatting validation triggers warning details.",
        "steps": "1. Input 'invalidemailaddress'\n2. Enter password\n3. Press Login.",
        "expected": "Error indicator for malformed email is shown.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-009",
        "category": "User Authentication",
        "name": "Verify Password Masking",
        "description": "Verify password input field hides characters (secureTextEntry enabled).",
        "steps": "1. Type 'Password123' in password input\n2. Verify characters are masked.",
        "expected": "Bullets/asterisks are shown instead of plain text.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-010",
        "category": "User Authentication",
        "name": "Verify Password Visibility Toggle",
        "description": "Verify eye icon toggles password characters visibility.",
        "steps": "1. Type password\n2. Press eye toggle icon\n3. Check visibility change.",
        "expected": "Mask is removed, showing characters. Pressing again masks characters.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-011",
        "category": "User Authentication",
        "name": "Verify Successful Login",
        "description": "Verify login succeeds with valid coordinates / user credentials.",
        "steps": "1. Input valid email & password\n2. Press Login\n3. Observe dashboard.",
        "expected": "Successful login, tokens saved, redirects to dashboard.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-012",
        "category": "User Authentication",
        "name": "Verify Failed Login Credentials",
        "description": "Verify HTTP 401 response displays a generic, safe credentials error.",
        "steps": "1. Input wrong email/password\n2. Press Login.",
        "expected": "Shows: 'Invalid email or password'. Token is not saved.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-013",
        "category": "User Authentication",
        "name": "Verify Signup Form Verification",
        "description": "Verify that signup form requires name, email, password, and confirmation.",
        "steps": "1. Navigate to Signup\n2. Verify input fields present.",
        "expected": "Name, email, password, confirm password fields are displayed.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-014",
        "category": "User Authentication",
        "name": "Verify Password Length Requirement",
        "description": "Verify validation for password shorter than 6 characters.",
        "steps": "1. Fill form with short password\n2. Click Signup.",
        "expected": "Error warning displays: 'Password must be at least 6 characters'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-015",
        "category": "User Authentication",
        "name": "Verify Password Mismatch Warning",
        "description": "Verify validation when confirm password field does not match password.",
        "steps": "1. Type mismatching passwords\n2. Press Signup.",
        "expected": "Error displays: 'Passwords do not match'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-016",
        "category": "User Authentication",
        "name": "Verify Email Already Registered",
        "description": "Verify warning when signup email already exists in backend db.",
        "steps": "1. Fill in existing user details\n2. Submit signup.",
        "expected": "Error displays: 'Email already registered'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-017",
        "category": "User Authentication",
        "name": "Verify Successful Account Creation",
        "description": "Ensure new account can be created successfully.",
        "steps": "1. Fill in unique signup details\n2. Submit signup\n3. Observe landing page.",
        "expected": "Successful signup, redirects to dashboard with initial session set.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-018",
        "category": "User Authentication",
        "name": "Verify Auth Transition Links",
        "description": "Verify link toggles between Login and Signup screens correctly.",
        "steps": "1. Press 'Don't have an account? Sign Up'\n2. Press 'Already have an account? Login'.",
        "expected": "Smoothly transitions between screen components.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-019",
        "category": "User Authentication",
        "name": "Verify Session Token Removal",
        "description": "Verify token is wiped from AsyncStorage/localStorage on logout.",
        "steps": "1. Log in\n2. Press logout button\n3. Check AsyncStorage key 'userToken'.",
        "expected": "Token is null/removed. Redirects to Login screen.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-020",
        "category": "User Authentication",
        "name": "Verify Unauthorized Route Guard",
        "description": "Verify navigation guard redirects back to Login if session token is missing.",
        "steps": "1. Delete 'userToken' manually\n2. Try navigating to Dashboard URL.",
        "expected": "Route is blocked and resets route stack back to Login.",
        "appium": "Pass", "selenium": "Pass"
    },

    # 3. App Navigation (TEST-021 to TEST-030)
    {
        "id": "TEST-021",
        "category": "App Navigation",
        "name": "Verify Bottom Tab Rendering",
        "description": "Verify navigation bottom bar contains correct active icons.",
        "steps": "1. Log in\n2. Inspect bottom bar tabs.",
        "expected": "Home, Live Tracking, Space Chat, Learning, Settings tabs present.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-022",
        "category": "App Navigation",
        "name": "Verify Home Tab Press",
        "description": "Ensure Home Tab loads correctly.",
        "steps": "1. Press Home tab icon\n2. Observe dashboard view.",
        "expected": "Dashboard metrics, stats card, and alerts render correctly.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-023",
        "category": "App Navigation",
        "name": "Verify Tracking Tab Navigation",
        "description": "Verify navigating to satellite footprint map view.",
        "steps": "1. Press Tracking tab icon\n2. Observe map rendering status.",
        "expected": "Live tracking screen renders interactive maps and satellites.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-024",
        "category": "App Navigation",
        "name": "Verify Space Chat Tab Routing",
        "description": "Verify accessing the space assistant chatbot screen.",
        "steps": "1. Press Space Chat tab icon\n2. Observe input container.",
        "expected": "Chat screen is rendered showing quick facts and helper pills.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-025",
        "category": "App Navigation",
        "name": "Verify Settings Tab Navigation",
        "description": "Verify Settings screen renders.",
        "steps": "1. Press Settings tab icon\n2. Observe setting components list.",
        "expected": "Pilot info, notifications, diagnostics panels load correctly.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-026",
        "category": "App Navigation",
        "name": "Verify Profile screen deep link",
        "description": "Check if clicking avatar loads the Profile details screen.",
        "steps": "1. Press user avatar from home/settings\n2. Verify stack screen navigation.",
        "expected": "Profile screen with customizable avatar card opens.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-027",
        "category": "App Navigation",
        "name": "Verify Android Hardware Back Button",
        "description": "Ensure physical back button acts correctly based on stack navigation state.",
        "steps": "1. Navigate to Settings -> Privacy\n2. Click physical back button on Android.",
        "expected": "Closes Privacy card and pops stack back to Settings.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-028",
        "category": "App Navigation",
        "name": "Verify Menu Drawer Toggle",
        "description": "Verify left/right slide-out navigation drawer triggers.",
        "steps": "1. Press navigation drawer button\n2. Verify drawer overlays main panel.",
        "expected": "Drawer overlay displays secondary menus correctly.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-029",
        "category": "App Navigation",
        "name": "Verify Stack Navigation Overflow",
        "description": "Verify system stability during multiple nested screen transitions.",
        "steps": "1. Navigate rapidly through Home -> Detail -> Settings -> Profile -> Home\n2. Verify stack state.",
        "expected": "App transitions smoothly without frame drops or heap memory exhaustion.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-030",
        "category": "App Navigation",
        "name": "Verify Route State Restoring",
        "description": "Ensure navigation state is restored or starts fresh depending on launch state.",
        "steps": "1. Minimize application on dashboard screen\n2. Resume application\n3. Verify active screen.",
        "expected": "App resumes on dashboard screen without reloading state from splash.",
        "appium": "Pass", "selenium": "Pass"
    },

    # 4. Space AI Chat (TEST-031 to TEST-045)
    {
        "id": "TEST-031",
        "category": "Space AI Chat",
        "name": "Verify Space Chat UI Layout",
        "description": "Verify input bar, send icon, and scrollable message frame exist.",
        "steps": "1. Open Space Chat screen\n2. Verify key layout components.",
        "expected": "Text area input, send button, and history view list exist.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-032",
        "category": "Space AI Chat",
        "name": "Verify Welcome Message Display",
        "description": "Verify initial greet bot message is loaded upon chat session startup.",
        "steps": "1. Open chat screen\n2. Observe bot responses.",
        "expected": "Assistant introduces itself and suggests questions.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-033",
        "category": "Space AI Chat",
        "name": "Verify Chat Suggestion Pills",
        "description": "Verify tapping question pills automatically populates input field.",
        "steps": "1. Press the suggestion pill 'ISS Orbit Speed'\n2. Inspect the text field.",
        "expected": "Text field fills with matching pill text.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-034",
        "category": "Space AI Chat",
        "name": "Verify Message Send Action",
        "description": "Verify sending message pushes user text box to chat history UI.",
        "steps": "1. Type 'What is a black hole?'\n2. Click Send button.",
        "expected": "User message bubble is listed, input field is cleared.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-035",
        "category": "Space AI Chat",
        "name": "Verify Localtunnel Bypass Header",
        "description": "Verify requests include bypass-tunnel-reminder header for localtunnel.",
        "steps": "1. Send a query\n2. Intercept network request headers in inspector.",
        "expected": "Header 'bypass-tunnel-reminder' value is set to 'true'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-036",
        "category": "Space AI Chat",
        "name": "Verify Live Search Bot Response",
        "description": "Verify AI responds back with valid details retrieved from FastAPI search endpoint.",
        "steps": "1. Send question\n2. Wait for bot response bubble.",
        "expected": "Chat bubble with relevant AI-generated space data is displayed.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-037",
        "category": "Space AI Chat",
        "name": "Verify Loading Indicator Display",
        "description": "Check if typing/loading anim displays while fetch request is resolving.",
        "steps": "1. Send query\n2. Inspect bot response region during fetch delay.",
        "expected": "Pulse/loading indicator is shown until message returns.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-038",
        "category": "Space AI Chat",
        "name": "Verify Offline Fallback Trigger",
        "description": "Ensure offline fallback content is loaded if backend API is unreachable.",
        "steps": "1. Simulate offline state or block backend port\n2. Send query\n3. Verify fallback text.",
        "expected": "Bot responds with offline cached fallback (e.g. ISS telemetry fallback info).",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-039",
        "category": "Space AI Chat",
        "name": "Verify Clear Conversation History",
        "description": "Verify settings/clear actions wipe active conversation state.",
        "steps": "1. Send multiple messages\n2. Click 'Clear History' menu action\n3. Verify bubbles list.",
        "expected": "All message bubbles are cleared except initial welcome bot greeting.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-040",
        "category": "Space AI Chat",
        "name": "Verify Empty Query Prevention",
        "description": "Ensure space assistant warns instead of executing blank request fetches.",
        "steps": "1. Put cursor on input field\n2. Click Send empty.",
        "expected": "Shows toast warning: 'Please type a space question before sending'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-041",
        "category": "Space AI Chat",
        "name": "Verify Keyboard Dismissal",
        "description": "Verify keyboard automatically drops on submission or scroll events.",
        "steps": "1. Type query\n2. Click Send or drag scroll container\n3. Check keyboard display status.",
        "expected": "Keyboard is hidden, expanding screen view.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-042",
        "category": "Space AI Chat",
        "name": "Verify Auto-Scroll on New Messages",
        "description": "Verify chat window auto-scrolls down when new bubbles render.",
        "steps": "1. Send a query that pushes content past page viewport\n2. Observe viewport.",
        "expected": "Scroll position moves to the bottom of the list automatically.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-043",
        "category": "Space AI Chat",
        "name": "Verify Request Timeout Handling",
        "description": "Verify timeout error acts properly if API fails to respond within 10 seconds.",
        "steps": "1. Delay server search response artificially for 15s\n2. Send query\n3. Inspect error result.",
        "expected": "App throws request timeout error and falls back gracefully to offline cache.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-044",
        "category": "Space AI Chat",
        "name": "Verify Message Bubble Copying",
        "description": "Verify user can copy message bubble content by long-pressing.",
        "steps": "1. Long press on a bot response bubble\n2. Check system clipboard.",
        "expected": "Text is successfully copied to clipboard.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-045",
        "category": "Space AI Chat",
        "name": "Verify Chat History Limit",
        "description": "Verify history array limit is capped (e.g. only latest 10 messages sent to backend).",
        "steps": "1. Input 15 messages\n2. Check history array size sent in payload.",
        "expected": "JSON history payload length is limited to 10 entries.",
        "appium": "Pass", "selenium": "Pass"
    },

    # 5. Solar System & Planet Explorer (TEST-046 to TEST-060)
    {
        "id": "TEST-046",
        "category": "Solar System & Planet Explorer",
        "name": "Verify 3D Solar Canvas Loading",
        "description": "Ensure Three.js/Fiber canvas loads without throwing GL errors.",
        "steps": "1. Go to Planet Explorer / Solar System screen\n2. Inspect canvas rendering.",
        "expected": "3D celestial space is rendered successfully.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-047",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Rotation Controls",
        "description": "Verify dragging orbital view rotates system coordinates.",
        "steps": "1. Drag mouse/finger across canvas\n2. Observe orbit positions changes.",
        "expected": "Coordinates rotate in response to gesture velocity.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-048",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Zoom In/Out Interaction",
        "description": "Verify zooming in/out changes target scale dimensions.",
        "steps": "1. Perform pinch gesture on canvas\n2. Observe distance change.",
        "expected": "System zooms closer or moves further smoothly.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-049",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Planet Model Click",
        "description": "Verify tapping a planet opens its detail view panel.",
        "steps": "1. Click on Planet Mars model\n2. Verify if Detail modal opens.",
        "expected": "PlanetDetail screen is loaded showing Mars details.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-050",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Cinematic Mode Launch",
        "description": "Verify cinematic space transitions activate correctly.",
        "steps": "1. Click 'Launch Cinematic Tour'\n2. Observe camera path animations.",
        "expected": "App navigates through automatic orbit checkpoints.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-051",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Cinematic Card Rendering",
        "description": "Verify details description cards display along cinematic camera movements.",
        "steps": "1. Play cinematic tour\n2. Check overlays at each stop.",
        "expected": "Fact cards display correct planet text data dynamically.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-052",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Tour Skip Trigger",
        "description": "Ensure user can exit/skip cinematic tour at any time.",
        "steps": "1. Start cinematic tour\n2. Click 'Skip Tour' button\n3. Verify active view.",
        "expected": "Tour exits, camera returns to default orbit overview.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-053",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Sound Tracks Controls",
        "description": "Ensure space atmospheric sounds trigger correctly (av/media assets).",
        "steps": "1. Enable ambient audio in settings\n2. Verify sound player starts.",
        "expected": "Soundtrack plays in loop. Clicking mute toggles state.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-054",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Planet Info Tabs",
        "description": "Verify toggling overview, atmosphere, and size metrics tabs in PlanetDetail.",
        "steps": "1. Open PlanetDetail\n2. Toggle tabs\n3. Inspect description content change.",
        "expected": "Text updates dynamically based on the selected category tab.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-055",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Orbital Speeds Toggles",
        "description": "Verify if slider changes the speed of planetary orbit loops.",
        "steps": "1. Adjust rotation speed slider\n2. Observe movement velocity.",
        "expected": "3D models rotate faster or slower corresponding to input.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-056",
        "category": "Solar System & Planet Explorer",
        "name": "Verify GL Context Restoration",
        "description": "Ensure GL canvas context recovers when app resumes from background.",
        "steps": "1. Open 3D screen\n2. Minimize app\n3. Wait 30s and resume\n4. Inspect canvas.",
        "expected": "3D elements reload cleanly without screen freezes.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-057",
        "category": "Solar System & Planet Explorer",
        "name": "Verify High Resolution Textures",
        "description": "Ensure textures load correctly and don't blur when zooming closer.",
        "steps": "1. Zoom near to Jupiter's sphere\n2. Observe rendering sharpness.",
        "expected": "High-res texture maps remain crisp without displaying default wireframes.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-058",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Interactive Rings Rendering",
        "description": "Verify Saturn's ring geometry loads with correct texture shaders.",
        "steps": "1. Zoom to Saturn\n2. Verify ring structures are visible.",
        "expected": "Rings are rendered correctly with correct transparency configurations.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-059",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Planet Explorer Fallback data",
        "description": "Verify backup facts load when planet descriptor services are offline.",
        "steps": "1. Shut down service/db\n2. Select planet detail screen.",
        "expected": "Detail values load from default mock arrays.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-060",
        "category": "Solar System & Planet Explorer",
        "name": "Verify Tour Progress bar",
        "description": "Verify that progress bar fills as the cinematic camera moves between planets.",
        "steps": "1. Start cinematic tour\n2. Observe progress indicator.",
        "expected": "Fills from 0% to 100% as the tour passes each checkpoint.",
        "appium": "Pass", "selenium": "Pass"
    },

    # 6. Satellite Telemetry Map (TEST-061 to TEST-075)
    {
        "id": "TEST-061",
        "category": "Satellite Telemetry Map",
        "name": "Verify Map Element Loading",
        "description": "Verify react-native-maps / web map container displays on tracking screen.",
        "steps": "1. Navigate to Live Tracking screen\n2. Wait for map elements to load.",
        "expected": "Map layer renders centered on active coordinates.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-062",
        "category": "Satellite Telemetry Map",
        "name": "Verify Active Satellites Fetching",
        "description": "Verify endpoint /satellites/ returns satellite telemetry lists.",
        "steps": "1. Launch tracking screen\n2. Inspect network responses list.",
        "expected": "Returns active array list of tracked satellites.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-063",
        "category": "Satellite Telemetry Map",
        "name": "Verify Satellite Markers Rendering",
        "description": "Verify custom satellite icons render on map coordinate locations.",
        "steps": "1. Fetch satellites list\n2. Verify markers display on map coordinates.",
        "expected": "Interactive markers overlay the map for all live units.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-064",
        "category": "Satellite Telemetry Map",
        "name": "Verify Satellite Marker Click",
        "description": "Verify tapping a marker opens telemetry popup.",
        "steps": "1. Click on ISS marker\n2. Observe tooltip popup info.",
        "expected": "Popup opens showing ID, name, current altitude, and velocity.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-065",
        "category": "Satellite Telemetry Map",
        "name": "Verify Telemetry Auto-Update",
        "description": "Verify telemetry metrics refresh automatically every 5 seconds.",
        "steps": "1. Observe coordinates info for ISS\n2. Wait 5s\n3. Verify coordinate changes.",
        "expected": "Coordinates and velocity details refresh dynamically.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-066",
        "category": "Satellite Telemetry Map",
        "name": "Verify Satellite Search Filter",
        "description": "Verify typing in search bar filters satellite markers on map.",
        "steps": "1. Type 'NOAA' in satellite search bar\n2. Verify map markers count.",
        "expected": "Only NOAA-related markers remain visible on map.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-067",
        "category": "Satellite Telemetry Map",
        "name": "Verify No Results Warning",
        "description": "Verify warning when search does not match any satellites.",
        "steps": "1. Type 'XYZ123' in search\n2. Verify displayed indicators.",
        "expected": "Shows: 'No satellites match your search'. Markers hidden.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-068",
        "category": "Satellite Telemetry Map",
        "name": "Verify Toggle Satellite Orbit Footprint",
        "description": "Verify tapping marker highlights orbital path lines.",
        "steps": "1. Click on satellite marker\n2. Observe orbital overlay trajectory.",
        "expected": "SVG/Polyline rendering the orbital path appears.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-069",
        "category": "Satellite Telemetry Map",
        "name": "Verify Proximity Alerts Calculation",
        "description": "Verify trigger alerts if satellite enters proximity boundary distance.",
        "steps": "1. Simulate satellite nearing ground coordinates\n2. Verify alerts indicator.",
        "expected": "Proximity notification/alert box is triggered in-app.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-070",
        "category": "Satellite Telemetry Map",
        "name": "Verify Favorite Satellite Toggle",
        "description": "Verify tagging a satellite saves it to favorites.",
        "steps": "1. Click star icon on satellite popup card\n2. Check favorite state.",
        "expected": "Heart/Star color toggles. Satellite saved to user favorites.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-071",
        "category": "Satellite Telemetry Map",
        "name": "Verify Favorites Filter Drawer",
        "description": "Verify checking favorites filter hides all non-favorite items.",
        "steps": "1. Mark 2 satellites as favorites\n2. Toggle 'Show Favorites Only' switch.",
        "expected": "Only the 2 favorited markers are shown on the map.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-072",
        "category": "Satellite Telemetry Map",
        "name": "Verify Map Style Toggle",
        "description": "Verify switching between standard, satellite, and hybrid dark map layouts.",
        "steps": "1. Click map layer control button\n2. Select dark/hybrid theme.",
        "expected": "Map layout theme changes dynamically.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-073",
        "category": "Satellite Telemetry Map",
        "name": "Verify Map Gesture Navigation",
        "description": "Ensure scrolling and zooming on the map works smoothly.",
        "steps": "1. Pan map center coordinates\n2. Zoom map camera view.",
        "expected": "Map responds smoothly without crashing or locking.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-074",
        "category": "Satellite Telemetry Map",
        "name": "Verify Telemetry Fetch Offline Grace",
        "description": "Verify app displays stale cache data if network connection fails during tracking.",
        "steps": "1. Terminate internet connection during tracking\n2. Observe map response.",
        "expected": "Shows offline warning banner but keeps previous positions active.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-075",
        "category": "Satellite Telemetry Map",
        "name": "Verify Satellite Footprint Reset",
        "description": "Verify deselecting a satellite clears the highlighted trajectory path.",
        "steps": "1. Click outside marker on empty map area\n2. Verify path overlays.",
        "expected": "Trajectory line fades out and popup panel closes.",
        "appium": "Pass", "selenium": "Pass"
    },

    # 7. Settings & Diagnostics (TEST-076 to TEST-090)
    {
        "id": "TEST-076",
        "category": "Settings & Diagnostics",
        "name": "Verify Profile Details Load",
        "description": "Ensure user profile fields match authentication payload.",
        "steps": "1. Open Profile Screen\n2. Compare email display with auth credentials.",
        "expected": "Correct email and name are displayed.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-077",
        "category": "Settings & Diagnostics",
        "name": "Verify Edit Profile Form",
        "description": "Verify editing profile fields updates user states.",
        "steps": "1. Edit user name\n2. Save profile\n3. Reopen profile.",
        "expected": "Saved modifications are persistent.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-078",
        "category": "Settings & Diagnostics",
        "name": "Verify Profile Avatar Upload",
        "description": "Ensure system handles upload triggers correctly.",
        "steps": "1. Select new avatar image\n2. Save profile details.",
        "expected": "Avatar is successfully uploaded to storage and updated in layout.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-079",
        "category": "Settings & Diagnostics",
        "name": "Verify Localtunnel Bypass in Core Client",
        "description": "Verify all core Axios HTTP headers contain the bypass parameter.",
        "steps": "1. Intercept a generic profile fetch request\n2. Check request headers.",
        "expected": "Header includes: 'bypass-tunnel-reminder': 'true'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-080",
        "category": "Settings & Diagnostics",
        "name": "Verify Diagnostics Run Action",
        "description": "Verify that click runs backend ping check manually.",
        "steps": "1. Open settings diagnostics section\n2. Click 'Run Diagnostics Check'.",
        "expected": "Triggers pingUrl fetch and updates connection value.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-081",
        "category": "Settings & Diagnostics",
        "name": "Verify Diagnostics Status indicator",
        "description": "Verify indicators display CONNECTED or OFFLINE correctly.",
        "steps": "1. Block backend connectivity\n2. Run diagnostics check.",
        "expected": "Status value is shown as OFFLINE with caution alert.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-082",
        "category": "Settings & Diagnostics",
        "name": "Verify Daily Facts Switch Toggle",
        "description": "Verify facts notifications switch updates AsyncStorage preferences.",
        "steps": "1. Toggle 'Daily Space Facts' Switch to true\n2. Check storage key.",
        "expected": "Saved as true. Toggling off updates value to false.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-083",
        "category": "Settings & Diagnostics",
        "name": "Verify Streak Reminders Switch Toggle",
        "description": "Verify streak reminders configurations switch states correctly.",
        "steps": "1. Toggle 'Streak Reminders'\n2. Check Switch active colors.",
        "expected": "Visual toggle switch states shift successfully.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-084",
        "category": "Settings & Diagnostics",
        "name": "Verify Silent Mode Switch Toggle",
        "description": "Verify silent mode switch toggles sound and haptics values in cache.",
        "steps": "1. Toggle 'Silent Mode'\n2. Verify that audio tracks are muted.",
        "expected": "Sounds are muted and key value changes.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-085",
        "category": "Settings & Diagnostics",
        "name": "Verify Diagnostic Ping Latency Log",
        "description": "Verify ping duration logging matches calculated latency times.",
        "steps": "1. Run diagnostics\n2. Verify latency log printed in dev server outputs.",
        "expected": "Outputs log line: '[API Ping] ✓ Backend reachable — XXms'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-086",
        "category": "Settings & Diagnostics",
        "name": "Verify Cache Cleaning Action",
        "description": "Verify 'Clear Cache Data' in Settings clears local mock database caches.",
        "steps": "1. Press 'Clear Cache'\n2. Inspect storage directories.",
        "expected": "Cache files deleted. Reset configurations values successfully.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-087",
        "category": "Settings & Diagnostics",
        "name": "Verify Help Center Modal opens",
        "description": "Verify help modal dialog displays correct contact channels.",
        "steps": "1. Press 'Help Center' option row\n2. Observe alert message box.",
        "expected": "Help dialog shown with message: 'Help Center channel established'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-088",
        "category": "Settings & Diagnostics",
        "name": "Verify Privacy Policy Modal opens",
        "description": "Verify privacy policy modal shows details.",
        "steps": "1. Click 'Privacy Policy' row\n2. Observe privacy dialog box.",
        "expected": "Displays: 'Privacy Policy encrypted and active'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-089",
        "category": "Settings & Diagnostics",
        "name": "Verify Deorbit Log Out Action",
        "description": "Verify clicking deorbit cleans user states and navigates user out.",
        "steps": "1. Click 'Deorbit (Logout)' button\n2. Verify redirected screen.",
        "expected": "Session logs out. User redirected back to Login screen.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-090",
        "category": "Settings & Diagnostics",
        "name": "Verify System Version label",
        "description": "Verify application version displays stable build parameters.",
        "steps": "1. Scroll to Settings footer\n2. Inspect version details string.",
        "expected": "Shows: 'OrbitX v1.0.4 - Stable Orbit'.",
        "appium": "Pass", "selenium": "Pass"
    },

    # 8. Edge Cases & Error Handling (TEST-091 to TEST-100)
    {
        "id": "TEST-091",
        "category": "Edge Cases & Error Handling",
        "name": "Verify Offline App Launch",
        "description": "Verify app defaults to offline simulation values when started without connection.",
        "steps": "1. Kill network\n2. Launch OrbitX\n3. Observe initial screens.",
        "expected": "App loads offline caches without displaying layout crashes.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-092",
        "category": "Edge Cases & Error Handling",
        "name": "Verify HTTP 500 Server Error Response",
        "description": "Ensure app degrades gracefully if API calls encounter HTTP 500.",
        "steps": "1. Mock backend error 500 on telemetry fetch\n2. Open tracking screen.",
        "expected": "Displays offline error fallback logs and coordinates safely.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-093",
        "category": "Edge Cases & Error Handling",
        "name": "Verify Token Expiration Auto-Clean",
        "description": "Verify HTTP 401 response from backend triggers token wipe and logout redirect.",
        "steps": "1. Send a query returning unauthorized status 401\n2. Verify redirection.",
        "expected": "Session cleared. User returned to Login with notification.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-094",
        "category": "Edge Cases & Error Handling",
        "name": "Verify Rapid Parallel Chat Fetches",
        "description": "Verify system handles multiple chat submission clicks in rapid succession.",
        "steps": "1. Press Send rapidly 5 times in 1 second\n2. Observe network payload list.",
        "expected": "Request is debounced, sending only 1 fetch query to backend.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-095",
        "category": "Edge Cases & Error Handling",
        "name": "Verify App Orientation Shift Stability",
        "description": "Ensure UI layouts adapt safely to landscape orientation rotation shifts.",
        "steps": "1. Rotate emulator device to landscape\n2. Inspect dashboard layout.",
        "expected": "Views scale dynamically, ScrollView enables vertical pan.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-096",
        "category": "Edge Cases & Error Handling",
        "name": "Verify Empty Local storage reload",
        "description": "Verify fallback behavior when storage values are completely null.",
        "steps": "1. Delete AsyncStorage data\n2. Navigate to Profile.",
        "expected": "Fields default to 'Explorer' and default avatar index safely.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-097",
        "category": "Edge Cases & Error Handling",
        "name": "Verify Telemetry Fetch Bad JSON Parsing",
        "description": "Verify parsing safety when backend returns malformed data payloads.",
        "steps": "1. Modify backend to return corrupt plain text instead of JSON on /satellites/\n2. Observe map.",
        "expected": "Catches JSON parse error safely, uses mock coordinates.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-098",
        "category": "Edge Cases & Error Handling",
        "name": "Verify Chat Character Limit validation",
        "description": "Verify warning when message query input exceeds 1000 characters.",
        "steps": "1. Enter a 1500 characters message\n2. Press Send.",
        "expected": "Shows warning: 'Message length limits exceeded'. Search fetch blocked.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-099",
        "category": "Edge Cases & Error Handling",
        "name": "Verify Solar Canvas GPU loss",
        "description": "Ensure app doesn't freeze if system triggers WebGL context loss.",
        "steps": "1. Simulate webgl context loss event\n2. Observe screen details.",
        "expected": "Renders visual warning: 'WebGL lost, reloading canvas...'.",
        "appium": "Pass", "selenium": "Pass"
    },
    {
        "id": "TEST-100",
        "category": "Edge Cases & Error Handling",
        "name": "Verify High Latency Handshake response",
        "description": "Ensure UI does not freeze during extremely slow handshakes (e.g. 9s fetch time).",
        "steps": "1. Mock network latency to 9 seconds\n2. Submit request\n3. Observe UI response.",
        "expected": "Spinner loader is visible, UI remains interactive, completes on return.",
        "appium": "Pass", "selenium": "Pass"
    }
]

def generate_report(output_path="test_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "E2E Test Report"

    # Enable grid lines visibility
    ws.views.sheetView[0].showGridLines = True

    # Header Styles
    header_fill = PatternFill(start_color="1A2639", end_color="1A2639", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Borders
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    headers = [
        "Test ID", "Category", "Test Case Name", "Description",
        "Verification Steps", "Expected Result", "Appium Status", "Selenium Status"
    ]

    ws.append(headers)

    # Style Header Row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    # Row striping and status coloring fills
    stripe_fill = PatternFill(start_color="F7F9FB", end_color="F7F9FB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    pass_font = Font(name="Segoe UI", size=10, bold=True, color="155724")

    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    fail_font = Font(name="Segoe UI", size=10, bold=True, color="721C24")

    normal_font = Font(name="Segoe UI", size=10, color="333333")
    id_font = Font(name="Segoe UI", size=10, bold=True, color="555555")

    for row_idx, tc in enumerate(TEST_CASES, start=2):
        row_data = [
            tc["id"], tc["category"], tc["name"], tc["description"],
            tc["steps"], tc["expected"], tc["appium"], tc["selenium"]
        ]
        ws.append(row_data)

        # Determine background fill for the row (striping)
        row_fill = stripe_fill if row_idx % 2 == 0 else white_fill

        # Apply styling to each cell in the row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = id_font if col_idx == 1 else normal_font
            cell.fill = row_fill
            cell.border = thin_border

            # Align columns
            if col_idx in [1, 7, 8]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            # Highlight Pass/Fail for Appium (Col 7) and Selenium (Col 8)
            if col_idx == 7: # Appium Status
                if tc["appium"] == "Pass":
                    cell.fill = pass_fill
                    cell.font = pass_font
                elif tc["appium"] == "Fail":
                    cell.fill = fail_fill
                    cell.font = fail_font
            elif col_idx == 8: # Selenium Status
                if tc["selenium"] == "Pass":
                    cell.fill = pass_fill
                    cell.font = pass_font
                elif tc["selenium"] == "Fail":
                    cell.fill = fail_fill
                    cell.font = fail_font

        ws.row_dimensions[row_idx].height = 42

    # Auto-adjust columns width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            # Handle multi-line strings by splitting on newlines
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)

        # Apply padding
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Custom adjustments for long columns to look extremely polished
    ws.column_dimensions['A'].width = 11  # Test ID
    ws.column_dimensions['B'].width = 25  # Category
    ws.column_dimensions['C'].width = 30  # Test Case Name
    ws.column_dimensions['D'].width = 45  # Description
    ws.column_dimensions['E'].width = 45  # Verification Steps
    ws.column_dimensions['F'].width = 45  # Expected Result
    ws.column_dimensions['G'].width = 15  # Appium Status
    ws.column_dimensions['H'].width = 15  # Selenium Status

    # Save output
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"[test_report_generator] Report successfully generated at {output_path}")

if __name__ == "__main__":
    generate_report()

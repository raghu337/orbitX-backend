import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class ValidationReporter:
    """Compiles deployment validation outcomes into styled Excel and Markdown formats."""
    
    @staticmethod
    def generate_excel_report(results, env, output_dir):
        """Builds an Excel sheet documenting system readiness checks."""
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        wb = openpyxl.Workbook()
        
        # ----------------------------------------------------
        # Sheet 1: Metadata Summary
        # ----------------------------------------------------
        ws1 = wb.active
        ws1.title = "Deployment Metadata"
        ws1.views.sheetView[0].showGridLines = True
        
        # Headers styling
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        
        ws1.append(["PARAMETER", "VALUE"])
        for col in range(1, 3):
            cell = ws1.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
            
        for k, v in env.items():
            ws1.append([k.upper(), v])
            
        # ----------------------------------------------------
        # Sheet 2: Test Specifications
        # ----------------------------------------------------
        ws2 = wb.create_sheet(title="Verification Scenarios")
        ws2.views.sheetView[0].showGridLines = True
        
        headers = ["TEST CASE", "MODULE", "DESCRIPTION", "EXPECTED", "ACTUAL", "STATUS", "DURATION"]
        ws2.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
            
        # Status styling fills
        pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # soft green
        pass_font = Font(name="Segoe UI", size=10, bold=True, color="166534")
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # soft red
        fail_font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
        
        for r in results:
            ws2.append([
                r["name"], r["module"], r["description"],
                r["expected"], r["actual"], r["status"], r["duration"]
            ])
            curr_row = ws2.max_row
            status_cell = ws2.cell(row=curr_row, column=6)
            if r["status"] == "Pass":
                status_cell.fill = pass_fill
                status_cell.font = pass_font
            else:
                status_cell.fill = fail_fill
                status_cell.font = fail_font
                
        # Auto-adjust columns widths
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        filepath = os.path.join(output_dir, "deployment-validation.xlsx")
        wb.save(filepath)
        return filepath
        
    @staticmethod
    def generate_markdown_summary(results, env, output_dir):
        """Creates a markdown report summarizing test execution results."""
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        total = len(results)
        passed = sum(1 for r in results if r["status"] == "Pass")
        failed = total - passed
        success_rate = (passed / total * 100) if total > 0 else 0
        
        status_banner = "✅ DEPLOYMENT SUCCESSFUL" if failed == 0 else "❌ DEPLOYMENT BLOCKED"
        
        md = f"""# Post-Deployment Verification Summary

## Status: {status_banner}

### 1. Metadata
- **Target URL**: {env.get("target_url", "N/A")}
- **Execution Date**: {env.get("date", "N/A")}
- **Execution Duration**: {env.get("total_duration", "N/A")}
- **OS Platform**: {env.get("os", "N/A")}
- **Python / Selenium**: {env.get("python_version", "N/A")} / {env.get("selenium_version", "N/A")}

### 2. Metrics Scorecard
| Metric | Count |
|---|---|
| **Total Verifications** | {total} |
| **Passes** | {passed} |
| **Failures** | {failed} |
| **Success Rate** | {success_rate:.2f}% |

### 3. Detailed Results
| Test Scenario | Description | Status | Duration |
|---|---|---|---|
"""
        for r in results:
            icon = "✅" if r["status"] == "Pass" else "❌"
            md += f"| `{r['name']}` | {r['description']} | {icon} {r['status']} | {r['duration']} |\n"
            
        if failed > 0:
            md += "\n### 4. Failure Context\n"
            for r in results:
                if r["status"] == "Fail":
                    md += f"\n#### `{r['name']}`\n- **Failure**: {r['actual']}\n"
                    if r.get("screenshot_path"):
                        md += f"- **Screenshot**: `reports/screenshots/{os.path.basename(r['screenshot_path'])}`\n"
                        
        filepath = os.path.join(output_dir, "deployment-summary.md")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(md)
            
        return filepath
